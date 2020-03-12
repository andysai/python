import openpyxl
from openpyxl.styles import Color, Font, Alignment

path = '../test/test_1/test.xlsx'

file = openpyxl.load_workbook(path)
sheet = file['19年10月']

reward_ws = file.create_sheet(title='19年11月')

# A1到H2列进行合并
reward_ws.merge_cells('A1:H2')
reward_ws.cell(1,1).value = '19年11月'

# A4和A5进行合并
reward_ws.merge_cells('A4:A5')
reward_ws.cell(4,1).value = '序号'
list = []
# 添加序号
for i in range(1,24):
    if i == 20:
        continue
    else:
        list.append(i)
print(list)
for k in list:
    if k >= 21:
        number = k + 4
    else:
        number = k + 5
    reward_ws.cell(number, 1).value = k

# B4和B5进行合并
reward_ws.merge_cells('B4:B5')
reward_ws.cell(4,2).value = '项目'



# C4和C5进行合并
reward_ws.merge_cells('C4:C5')
reward_ws.cell(4,3).value = '数量'

# D4和E4进行合并
reward_ws.merge_cells('D4:E4')
reward_ws.cell(4,4).value = '固定服务费'


reward_ws['D5'] = "当月运行天数"
reward_ws['E5'] = "固定服务费\n结算（元/柜）"

reward_ws.merge_cells('F4:G4')
reward_ws.cell(4,6).value = '浮动服务费'

reward_ws['F5'] = "机架超出电量(A)"
reward_ws['G5'] = "浮动服务费结算（元/A）"

reward_ws.merge_cells('H4:H5')
reward_ws.cell(4,8).value = '月服务费用（含税，单位：元）'

reward_ws['A28'] = '合计'

reward_ws.merge_cells('B28:G28')
reward_ws.cell(28,2).value = '月度结算费用（含税）'


reward_ws.merge_cells('A30:H31')
reward_ws.cell(30,1).value = '甲方签字盖章：                                乙方签字盖章：'
file.save('test_1/1.xlsx')