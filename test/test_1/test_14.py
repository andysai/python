# 导入模块
import openpyxl
# 打开表格
wb = openpyxl.load_workbook('新机房统计表.xlsx')
sheet = wb['机架服务器设备总表']
# 创建空列表
data = []
max_row = sheet.max_row
max_col = sheet.max_column
for i in sheet:
    data_1 = []
    for cell in i:
        data_1.append(cell.value)
    data.append(data_1)
#print(data)
#print(max_row,max_col)
physics_list = []
fictitious_list = []
reward_ws = wb.create_sheet(title='物理机')
reward_wd = wb.create_sheet(title='云主机')
for i in data:
    name = i[7] + '-' + i[8]
    CPU = i[15]
    Memory = i[16]
    Hard = i[17]
    Customer = i[33]
    fictitious = i[9]
    message = [name, CPU, Memory, Hard, Customer]
    if fictitious == '是':
        reward_wd.append(message)
    else:
        reward_ws.append(message)
#    print(name,CPU)
wb.save('test_1.xlsx')