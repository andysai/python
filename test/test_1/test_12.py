# 导入模块
import openpyxl
# 打开excle文档
wb = openpyxl.load_workbook(r'E:\20191202.xlsx')
# 当前活跃表单
table = wb.active
# 获取行数和列数
col_range = table['A:L']
row_range = table['1:123']
list = []
print('{}行 {}列'.format(table.max_row, table.max_column))
for row in row_range:
    values = []
    row = table.row_values(row)
    for col in col_range:
        values.append(col)
    list.append(values)
print(list)