import openpyxl,math
path = 'test.xlsx'

file = openpyxl.load_workbook(path)
sheet = file['19年10月']
sheet.title = '19年12月'
sheet.cell(1,1).value = '19年12月'
# 获取数量
first_list = []
first_column = sheet['C']
for i in range(5,len(first_column)-4):
    first_list.append(first_column[i].value)

# 获取当月运行天数
second_list = []
second_column = sheet['D']
for i in range(5,len(second_column)-4):
    second_list.append(second_column[i].value)

# 固定服务费结算（元/柜）
third_list = []
third_column = sheet['E']
for i in range(5,len(third_column)-4):
    third_list.append(third_column[i].value)

# 机架超出电量(A)
fourth_list = []
fourth_column = sheet['F']
for i in range(5,len(fourth_column)-4):
    fourth_list.append(fourth_column[i].value)

# 浮动服务费结算（元/A）
fivth_list = []
fivth_column = sheet['G']
for i in range(5,len(fivth_column)-4):
    fivth_list.append(fivth_column[i].value)

data = []
message = [first_list]+ [second_list] + [third_list] + [fourth_list] + [fivth_list]
data.append(message)

price_list = []
Electric_quantity = []
for Eq in data[0][3]:
    if Eq == None:
        Electric_quantity.append(0)
    else:
        Electric_quantity.append(Eq)
Floating_price = []

for Fp in data[0][4]:
    if Fp == None:
        Floating_price.append(0)
    else:
        Floating_price.append(Fp)
date = int(input('请输入当月的总天数:'))
for i in range(22):
    num = data[0][0][i]
    day = data[0][1][i]
    money = data[0][2][i]
    price = round(num * day / date * money + Electric_quantity[i] * Floating_price[i])
    price_list.append(price)

for q in range(len(price_list)):
    number = 6 + q
    sheet.cell(number,8).value = price_list[q]
sum = 0
for w in range(len(price_list)):
    sum = sum + price_list[w]
    w += 1
sheet.cell(28,8).value = sum
file.save('2019年12月.xlsx')

